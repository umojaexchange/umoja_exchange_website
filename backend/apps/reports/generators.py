"""
Report generators for PDF (reportlab) and Excel (openpyxl).
"""
from datetime import datetime
from decimal import Decimal
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import HRFlowable, Paragraph, SimpleDocTemplate, Table, TableStyle

YELLOW = colors.HexColor("#FACC15")
BLACK = colors.HexColor("#0F0F0F")
DARK_GRAY = colors.HexColor("#1C1C1C")
LIGHT_GRAY = colors.HexColor("#F3F4F6")
WHITE = colors.white


# ─── PDF ─────────────────────────────────────────────────────────────────────
def generate_pdf_report(report_type, queryset, date_from=None, date_to=None):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=1.5 * cm, leftMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title", parent=styles["Normal"], fontSize=18, fontName="Helvetica-Bold", textColor=BLACK, spaceAfter=6)
    sub_style = ParagraphStyle("Sub", parent=styles["Normal"], fontSize=10, textColor=colors.HexColor("#4B5563"), spaceAfter=12)

    elements = []

    # Header
    elements.append(Paragraph("UMOJA EXCHANGE", title_style))
    period = ""
    if date_from and date_to:
        period = f" | Period: {date_from} to {date_to}"
    elements.append(Paragraph(f"{report_type.upper()} REPORT{period} | Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", sub_style))
    elements.append(HRFlowable(width="100%", thickness=2, color=YELLOW, spaceAfter=12))

    if report_type == "purchases":
        _build_purchases_pdf(elements, queryset)
    elif report_type == "sales":
        _build_sales_pdf(elements, queryset)

    doc.build(elements)
    buffer.seek(0)
    return buffer


def _table_style(has_total=False):
    style = [
        ("BACKGROUND", (0, 0), (-1, 0), BLACK),
        ("TEXTCOLOR", (0, 0), (-1, 0), WHITE),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LIGHT_GRAY]),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
    ]
    if has_total:
        style += [
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FEF9C3")),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ]
    return TableStyle(style)


def _build_purchases_pdf(elements, qs):
    headers = ["#", "Supplier", "USDT", "Rate (TZS)", "Amount Paid (TZS)", "Method", "Date"]
    data = [headers]
    totals = {"usdt": Decimal(0), "paid": Decimal(0)}
    for i, p in enumerate(qs, 1):
        totals["usdt"] += p.usdt_amount
        totals["paid"] += p.amount_paid_tzs
        data.append([
            str(i), p.supplier_name, f"{p.usdt_amount:,.2f}",
            f"{p.rate_tzs:,.2f}", f"{p.amount_paid_tzs:,.2f}",
            p.get_payment_method_display(), p.created_at.strftime("%d/%m/%Y"),
        ])
    data.append(["", "TOTAL", f"{totals['usdt']:,.2f}", "", f"{totals['paid']:,.2f}", "", ""])
    t = Table(data, repeatRows=1)
    t.setStyle(_table_style(has_total=True))
    elements.append(t)


def _build_sales_pdf(elements, qs):
    headers = ["#", "Customer", "USDT", "Sale Rate", "Paid (TZS)", "Avg Buy", "Margin", "Profit (TZS)", "Method", "Date"]
    data = [headers]
    totals = {"usdt": Decimal(0), "paid": Decimal(0), "profit": Decimal(0)}
    for i, s in enumerate(qs, 1):
        totals["usdt"] += s.usdt_amount
        totals["paid"] += s.paid_amount_tzs
        totals["profit"] += s.profit_tzs
        data.append([
            str(i), s.customer_name, f"{s.usdt_amount:,.2f}",
            f"{s.sale_rate_tzs:,.2f}", f"{s.paid_amount_tzs:,.2f}",
            f"{s.avg_buy_rate:,.2f}", f"{s.profit_margin:,.2f}",
            f"{s.profit_tzs:,.2f}", s.get_payment_method_display(),
            s.created_at.strftime("%d/%m/%Y"),
        ])
    data.append(["", "TOTAL", f"{totals['usdt']:,.2f}", "", f"{totals['paid']:,.2f}", "", "", f"{totals['profit']:,.2f}", "", ""])
    t = Table(data, repeatRows=1)
    t.setStyle(_table_style(has_total=True))
    elements.append(t)


# ─── Excel ────────────────────────────────────────────────────────────────────
def generate_excel_report(report_type, queryset, date_from=None, date_to=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report_type.capitalize()

    # Styles
    black_fill = PatternFill(start_color="0F0F0F", end_color="0F0F0F", fill_type="solid")
    light_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    yellow_font = Font(name="Calibri", bold=True, color="0F0F0F", size=11)
    white_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    thin = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    # Title row
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"UMOJA EXCHANGE — {report_type.upper()} REPORT"
    title_cell.font = yellow_font
    title_cell.fill = black_fill
    title_cell.alignment = center
    ws.row_dimensions[1].height = 24

    # Period row
    ws.merge_cells("A2:J2")
    period_cell = ws["A2"]
    period_cell.value = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')} | Period: {date_from or 'All'} to {date_to or 'All'}"
    period_cell.font = Font(name="Calibri", size=9, color="4B5563")
    period_cell.alignment = center
    ws.row_dimensions[2].height = 16

    if report_type == "purchases":
        _build_purchases_excel(ws, queryset, black_fill, light_fill, header_font, white_font, thin, left, right)
    elif report_type == "sales":
        _build_sales_excel(ws, queryset, black_fill, light_fill, header_font, white_font, thin, left, right)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _apply_row(ws, row_num, values, is_header=False, is_total=False, black_fill=None, light_fill=None, header_font=None, white_font=None, thin=None, left=None, right=None):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.border = thin
        c.alignment = right if isinstance(val, int | float) else left
        if is_header:
            c.fill = black_fill
            c.font = header_font
        elif is_total:
            c.fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")
            c.font = Font(name="Calibri", bold=True, size=10)
        elif row_num % 2 == 0:
            c.fill = light_fill


def _build_purchases_excel(ws, qs, black_fill, light_fill, header_font, white_font, thin, left, right):
    headers = ["#", "Supplier", "USDT Amount", "Rate (TZS)", "Amount Paid (TZS)", "Payment Method", "Date"]
    _apply_row(ws, 3, headers, is_header=True, black_fill=black_fill, header_font=header_font, thin=thin, left=left, right=right)
    totals = {"usdt": 0, "paid": 0}
    for i, p in enumerate(qs, 1):
        totals["usdt"] += float(p.usdt_amount)
        totals["paid"] += float(p.amount_paid_tzs)
        _apply_row(ws, 3 + i, [i, p.supplier_name, float(p.usdt_amount), float(p.rate_tzs), float(p.amount_paid_tzs), p.get_payment_method_display(), p.created_at.strftime("%d/%m/%Y")], light_fill=light_fill, thin=thin, left=left, right=right)
    row = 3 + qs.count() + 1
    _apply_row(ws, row, ["", "TOTAL", totals["usdt"], "", totals["paid"], "", ""], is_total=True, thin=thin, left=left, right=right)
    col_widths = [6, 30, 16, 16, 20, 18, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_sales_excel(ws, qs, black_fill, light_fill, header_font, white_font, thin, left, right):
    headers = ["#", "Customer", "USDT", "Sale Rate", "Paid (TZS)", "Avg Buy Rate", "Margin (TZS)", "Profit (TZS)", "Method", "Date"]
    _apply_row(ws, 3, headers, is_header=True, black_fill=black_fill, header_font=header_font, thin=thin, left=left, right=right)
    totals = {"usdt": 0, "paid": 0, "profit": 0}
    for i, s in enumerate(qs, 1):
        totals["usdt"] += float(s.usdt_amount)
        totals["paid"] += float(s.paid_amount_tzs)
        totals["profit"] += float(s.profit_tzs)
        _apply_row(ws, 3 + i, [i, s.customer_name, float(s.usdt_amount), float(s.sale_rate_tzs), float(s.paid_amount_tzs), float(s.avg_buy_rate), float(s.profit_margin), float(s.profit_tzs), s.get_payment_method_display(), s.created_at.strftime("%d/%m/%Y")], light_fill=light_fill, thin=thin, left=left, right=right)
    row = 3 + qs.count() + 1
    _apply_row(ws, row, ["", "TOTAL", totals["usdt"], "", totals["paid"], "", "", totals["profit"], "", ""], is_total=True, thin=thin, left=left, right=right)
    col_widths = [5, 28, 14, 14, 18, 16, 16, 18, 16, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
